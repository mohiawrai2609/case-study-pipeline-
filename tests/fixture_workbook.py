#!/usr/bin/env python3
"""Build a synthetic role workbook and prove build/extract_excel.py handles it.

    python tests/fixture_workbook.py

The extractor is the only stage that touches a real workbook, so it is the one
stage that cannot be exercised by rebuilding existing data files. This fixture
covers the three things most recently changed there, each of which had shipped
wrong at least once:

  breadth       must come out a FRACTION of the role's tasks, never a raw count
  vendor order  VERIFIED/CORRECTED entities must outrank OUTSTANDING ones
  anchorRole    must be attached from data/_anchors.json, never typed by hand

The fixture is deliberately adversarial on the second point: the vendor covering
the MOST tasks is OUTSTANDING, so the ordering is only correct if verification
really does outrank breadth.
"""
import io
import json
import os
import subprocess
import sys
import tempfile

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

APS, HRF = 0.50, 0.40                      # 0.50 * (1 - 0.40) * 100 == 30.0
RPI = round(APS * (1 - HRF) * 100, 2)
ANCHOR_SOC = '15-1252.00'                  # present in data/_anchors.json

TASKS = [
    ('T1', 'Draft routine correspondence from templates', 0.80, 4.1),
    ('T2', 'Reconcile invoices against purchase orders', 0.70, 4.0),
    ('T3', 'Schedule and coordinate meetings across teams', 0.55, 3.6),
    ('T4', 'Resolve escalated exceptions with judgement', 0.30, 4.4),
    ('T5', 'Negotiate terms directly with counterparties', 0.20, 4.6),
]

VENDORS = {
    'Verified Broad Co':   ('VERIFIED',    [('T1', 0.72, 78), ('T2', 0.66, 75), ('T3', 0.61, 71)]),
    'Corrected Mid Co':    ('CORRECTED',   [('T1', 0.69, 70), ('T2', 0.58, 66)]),
    'Outstanding Wide Co': ('OUTSTANDING', [('T1', 0.80, 80), ('T2', 0.79, 79),
                                            ('T3', 0.77, 77), ('T4', 0.75, 75)]),
    'Outstanding Deep Co': ('OUTSTANDING', [('T1', 0.91, 84)]),
}


def rows(ws, header, data):
    ws.append(header)
    for r in data:
        ws.append(list(r))


def build(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Role')
    for k, v in [('title', 'Fixture Role'), ('soc_code', '99-0000.FX'),
                 ('anchor_soc_code', ANCHOR_SOC), ('aps', APS), ('hrf', HRF),
                 ('rpi_pct', RPI),
                 ('occupation_group', 'Office and Administrative Support'),
                 ('job_zone', 3), ('us_emp_k', 120.0), ('risk_band', 'Moderate'),
                 ('timeline', '3-7 years'),
                 ('source_type', 'fixture'), ('ajci_pct', 12.0),
                 ('cognitive_aps_pct', 55.0), ('physical_aps_pct', 20.0),
                 ('task_count', len(TASKS)), ('traditional_tasks', 3),
                 ('ai_augmented_tasks', 1), ('ai_created_tasks', 1)]:
        ws.append([k, v, ''])

    ws = wb.create_sheet('Readiness')
    for k, v in [('readiness_pct', 62.0), ('readiness_band', 'Emerging'),
                 ('tasks_covered', 4), ('tasks_total', len(TASKS)),
                 ('coverage_pct', 80.0), ('distinct_vendors', len(VENDORS)),
                 ('total_vendors', len(VENDORS)), ('replacement_tools', 2),
                 ('ecosystem', 'Fragmented')]:
        ws.append([k, v, ''])

    ws = wb.create_sheet('Narrative')
    ws.append(['role_summary',
               'A fixture occupation used only to exercise the extractor.', ''])
    ws.append(['verdict',
               'Fixture verdict: the role scores %.1f against an anchor of 13.4.' % RPI, ''])

    rows(wb.create_sheet('Tasks'),
         ['task_ref', 'task_text', 'task_aps', 'importance', 'task_type'],
         [(r, t, a, i, 'traditional') for r, t, a, i in TASKS])

    # T5 is deliberately UNSERVED: tasks_covered above says 4 of 5, and the
    # workflow assertSpine() halts if those two ever disagree.
    cov = []
    for ref, _t, _a, _i in TASKS:
        served = ref != 'T5'
        cov.append((ref,
                    1 if ref in ('T1', 'T2') else 0,
                    1 if served else 0,
                    'SERVED' if served else 'UNSERVED'))
    rows(wb.create_sheet('Coverage Matrix'),
         ['task_ref', 'replacement', 'augmentation', 'status'], cov)

    ev = []
    for vname, (_status, pairs) in VENDORS.items():
        for ref, c, t in pairs:
            ev.append((vname, ref, c, t,
                       'replacement' if c > 0.7 else 'augmentation',
                       'Verified' if c > 0.65 else 'Credible',
                       'Production'))
    rows(wb.create_sheet('Evidence'),
         ['vendor', 'task_ref', 'task_coverage', 'trust_score',
          'automation_mode', 'cert_tier', 'evidence_grade'], ev)

    rows(wb.create_sheet('Products'), ['vendor', 'product_name'],
         [(v, v.split()[0] + ' Suite') for v in VENDORS])
    rows(wb.create_sheet('Vendors'),
         ['canonical_name', 'status', 'hq', 'founded'],
         [(v, 'active', 'Somewhere', 2015) for v in VENDORS])
    rows(wb.create_sheet('Verification'), ['vendor', 'status', 'finding'],
         [(v, st, 'fixture') for v, (st, _f) in VENDORS.items()])

    wb.save(path)


def main():
    fails = []
    with tempfile.TemporaryDirectory() as tmp:
        xlsx = os.path.join(tmp, 'fixture.xlsx')
        out = os.path.join(tmp, 'fixture.json')
        build(xlsx)
        r = subprocess.run([sys.executable, 'build/extract_excel.py', xlsx, out,
                            '--issue', '999'], cwd=ROOT, capture_output=True, text=True)
        if r.returncode:
            print('EXTRACTOR FAILED\n' + (r.stderr or r.stdout).strip()[:1500])
            return 1
        with io.open(out, encoding='utf-8') as fh:
            d = json.load(fh)

    V = d['vendors']
    T = len(d['tasks'])
    print('extracted: %d tasks, %d vendor cards, %d in roster'
          % (T, len(V), len(d.get('vendorRoster') or [])))

    # 1. breadth is a fraction of the task count
    for v in V:
        if not 0 < v['breadth'] <= 1:
            fails.append('breadth not a fraction for %s: %s' % (v['name'], v['breadth']))
        exp = round(len(v.get('tasks') or []) / T, 4)
        if abs(v['breadth'] - exp) > 0.001:
            fails.append('breadth %s != tasks/%d = %s for %s'
                         % (v['breadth'], T, exp, v['name']))

    # 2. an unverified entity must not outrank a verified one
    rank = {'VERIFIED': 2, 'CORRECTED': 2}
    seq = [rank.get(v['verification'], 0) for v in V]
    if seq != sorted(seq, reverse=True):
        fails.append('verification not respected in ranking: %s'
                     % [(v['name'], v['verification']) for v in V])
    if V and V[0]['verification'] == 'OUTSTANDING':
        fails.append('an OUTSTANDING vendor leads the order: %s' % V[0]['name'])

    # 3. the anchor came from the shared table, with its sourced pair
    a = d.get('anchorRole')
    if not a or a.get('soc') != ANCHOR_SOC:
        fails.append('anchorRole not attached from _anchors.json: %s' % (a,))
    elif a.get('aps') is None or a.get('hrf') is None:
        fails.append('anchor %s lost its sourced aps/hrf pair: %s' % (ANCHOR_SOC, a))

    # 4. unserved tasks agree with the readiness count
    un = [t for t in d['tasks'] if str(t.get('status', '')).upper() == 'UNSERVED']
    cov = (d.get('readiness') or {}).get('tasksCovered')
    if cov is not None and T - len(un) != cov:
        fails.append('%d served but readiness.tasksCovered = %s' % (T - len(un), cov))

    for v in V:
        print('   %-24s%-13sbreadth=%.3f  depth=%.2f'
              % (v['name'][:22], v['verification'], v['breadth'], v['depth']))
    print('   anchorRole -> %s (aps=%s, hrf=%s)'
          % (a.get('title') if a else None,
             a.get('aps') if a else '?', a.get('hrf') if a else '?'))

    if fails:
        print('\nFAIL')
        for f in fails:
            print('  -', f)
        return 1
    print('\nPASS -- extractor handles a fresh workbook correctly.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
